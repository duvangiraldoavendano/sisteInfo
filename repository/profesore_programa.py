from openpyxl import load_workbook

def limpiar(tabla_profe_progra):
    anio_peri = 'a'
    pos_cell = 3
    wb = load_workbook(tabla_profe_progra)
    sheet = wb['profe_progra']
    cell = anio_peri + str(pos_cell)

    while (sheet[cell].value != None):
        sheet[cell] = ""
        pos_cell += 3
        cell = anio_peri + str(pos_cell)

    wb.save(tabla_profe_progra)

def estadistica_profesores(anio,semest,archivo_resul,tabla_profe_progra):
    anio_peri='a'
    pos_cell=3
    wb = load_workbook(tabla_profe_progra)
    sheet = wb['profe_progra']
    cell = anio_peri + str(pos_cell)
    x=sheet[cell]

    while(sheet[cell].value!=None):
        pos_cell+=3
        cell = anio_peri + str(pos_cell)

    cell = anio_peri + str(pos_cell)
    sheet[cell]=str(anio)+'-'+str(semest)
    wb.save(tabla_profe_progra)

    #Procedimiento para hacer el conteo de los profesores

    wb_profe_listado=load_workbook(archivo_resul)
    sheet_profe_listado=wb_profe_listado["profesores"]

    #Contadores de tiempo completo, medio tiempo y catedra
    conta_tc=0
    conta_mt=0
    conta_c=0

    #Contadores de niveles academicos
    conta_tc_doc=0
    conta_tc_maes=0
    conta_tc_espe=0
    conta_tc_pro=0
    conta_tc_tecnolo=0
    conta_tc_tecni=0

    conta_mt_doc = 0
    conta_mt_maes = 0
    conta_mt_espe = 0
    conta_mt_pro = 0
    conta_mt_tecnolo = 0
    conta_mt_tecni = 0

    conta_c_doc = 0
    conta_c_maes = 0
    conta_c_espe = 0
    conta_c_pro = 0
    conta_c_tecnolo = 0
    conta_c_tecni = 0

    pos=2
    cell_dedicacion='j'+str(pos)
    cell_nacademico='f'+str(pos)

    while(sheet_profe_listado[cell_dedicacion].value!=None):
        if(sheet_profe_listado[cell_dedicacion].value=='TIEMPO COMPLETO'):
            conta_tc+=1
            if(sheet_profe_listado[cell_nacademico].value=='DOCTORADO'):
                conta_tc_doc+=1
            elif(sheet_profe_listado[cell_nacademico].value=='MAESTRÍA'):
                conta_tc_maes+=1
            elif(sheet_profe_listado[cell_nacademico].value=='ESPECIALIZACIÓN'):
                conta_tc_espe+=1
            elif(sheet_profe_listado[cell_nacademico].value=='PREGRADO'):
                conta_tc_pro+=1
            elif(sheet_profe_listado[cell_nacademico].value=='TECNOLOGÍA'):
                conta_tc_tecnolo+=1
            elif(sheet_profe_listado[cell_nacademico].value=='TÉCNICO'):
                conta_tc_tecni+=1
            else:
                pass
        elif(sheet_profe_listado[cell_dedicacion].value=='MEDIO TIEMPO'):
            conta_mt+=1
            if (sheet_profe_listado[cell_nacademico].value == 'DOCTORADO'):
                conta_mt_doc += 1
            elif (sheet_profe_listado[cell_nacademico].value == 'MAESTRÍA'):
                conta_mt_maes += 1
            elif (sheet_profe_listado[cell_nacademico].value == 'ESPECIALIZACIÓN'):
                conta_mt_espe += 1
            elif (sheet_profe_listado[cell_nacademico].value == 'PREGRADO'):
                conta_mt_pro += 1
            elif (sheet_profe_listado[cell_nacademico].value == 'TECNOLOGÍA'):
                conta_mt_tecnolo += 1
            elif (sheet_profe_listado[cell_nacademico].value == 'TÉCNICO'):
                conta_mt_tecni += 1
            else:
                pass
        elif(sheet_profe_listado[cell_dedicacion].value=='CATEDRA'):
            conta_c+=1
            if (sheet_profe_listado[cell_nacademico].value == 'DOCTORADO'):
                conta_c_doc += 1
            elif (sheet_profe_listado[cell_nacademico].value == 'MAESTRÍA'):
                conta_c_maes += 1
            elif (sheet_profe_listado[cell_nacademico].value == 'ESPECIALIZACIÓN'):
                conta_c_espe += 1
            elif (sheet_profe_listado[cell_nacademico].value == 'PREGRADO'):
                conta_c_pro += 1
            elif (sheet_profe_listado[cell_nacademico].value == 'TECNOLOGÍA'):
                conta_c_tecnolo += 1
            elif (sheet_profe_listado[cell_nacademico].value == 'TÉCNICO'):
                conta_c_tecni += 1
            else:
                pass
        else:
            pass

        pos += 1
        cell_dedicacion = 'j' + str(pos)
        cell_nacademico = 'f' + str(pos)



    #Procedimiento para guardar las estadísticas(el conteo de los profesores)

    wb = load_workbook(tabla_profe_progra)
    sheet = wb['profe_progra']

    #Tiempo completo
    sheet['c'+str(pos_cell)]=conta_tc
    sheet['d' + str(pos_cell)] = conta_tc_doc
    sheet['e' + str(pos_cell)] = conta_tc_maes
    sheet['f' + str(pos_cell)] = conta_tc_espe
    sheet['g' + str(pos_cell)] = conta_tc_pro
    sheet['h' + str(pos_cell)] = conta_tc_tecnolo
    sheet['i' + str(pos_cell)] = conta_tc_tecni

    #Medio tiempo
    sheet['c' + str(pos_cell+1)] = conta_mt
    sheet['d' + str(pos_cell+1)] = conta_mt_doc
    sheet['e' + str(pos_cell+1)] = conta_mt_maes
    sheet['f' + str(pos_cell+1)] = conta_mt_espe
    sheet['g' + str(pos_cell+1)] = conta_mt_pro
    sheet['h' + str(pos_cell+1)] = conta_mt_tecnolo
    sheet['i' + str(pos_cell+1)] = conta_mt_tecni

    #Catedra
    sheet['c' + str(pos_cell + 2)] = conta_c
    sheet['d' + str(pos_cell + 2)] = conta_c_doc
    sheet['e' + str(pos_cell + 2)] = conta_c_maes
    sheet['f' + str(pos_cell + 2)] = conta_c_espe
    sheet['g' + str(pos_cell + 2)] = conta_c_pro
    sheet['h' + str(pos_cell + 2)] = conta_c_tecnolo
    sheet['i' + str(pos_cell + 2)] = conta_c_tecni

    wb.save(tabla_profe_progra)
