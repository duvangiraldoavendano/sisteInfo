from openpyxl import load_workbook

#Función que obtiene los profesores de ciencias politicas de catedra y los
#de derecho que son de cátedra y a su vez va al archivo de programación para
#mirar cuantas de las horas que aparecen en las horas de profesores, le corresponden a cada pregrado.

def discriminacion_horas_trabajo(cedula_dere, cedula_cp,prof_institu, profes_vincu, profes_cat, prof_institu_cp, profes_vincu_cp, profes_cat_cp,lineas_cedulas,excel_discriminacion,anio,semest):

    # -----------------------------------------------PROFESORES DE CIENCIAS POLÍTICAS------------------------------------------------------------

    #Procedimiento para obtener los nombres de los profesores de ciencias politicas que
    #pertenecen al instituto de ciencias politicas.
    wb_prof_institu = load_workbook(prof_institu)
    sheet_prof_institu = wb_prof_institu["reporte"]
    pos_prof_institu = 'd2'
    pos_num_prof_institu = 2
    lista_pares = []
    par_ced_nombre_doc = []
    lista_pares_nom_ced_prof_institu = []
    nombre_prof_institu = []
    while (sheet_prof_institu[pos_prof_institu].value != None):
        pos_nom_prof_institu = 'e' + str(pos_num_prof_institu)
        nombre_prof_institu.append(sheet_prof_institu[pos_nom_prof_institu].value)
        par_ced_nombre_doc_prof_institu = []
        index=0

        for ced_dere in cedula_cp:
            index+=1
            if (sheet_prof_institu[pos_prof_institu].value == ced_dere):
                par_ced_nombre_doc_prof_institu.append(ced_dere)
                pos_nom_prof_institu = 'e' + str(pos_num_prof_institu)
                par_ced_nombre_doc_prof_institu.append(sheet_prof_institu[pos_nom_prof_institu].value)
                ##pos_doc = 'e' + str(pos_num)
                # par_ced_nombre_doc.append("CC")
                break
        lista_pares_nom_ced_prof_institu.append(par_ced_nombre_doc_prof_institu)
        pos_num_prof_institu = pos_num_prof_institu + 1
        pos_prof_institu = 'd' + str(pos_num_prof_institu)

    nombre_prof_institu = set(nombre_prof_institu)

    #Procedimiento para eliminar de lista_pares_nom_ced_prof_institu las listas que sean vacías.
    resul = []
    for i in lista_pares_nom_ced_prof_institu:
        if i not in resul:
            resul.append(i)
    res = []
    for i in resul:
        if (i != []):
            res.append(i)

    lista_pares_nom_ced_prof_institu = res

    #Procedimiento para obtener los profesores de ciencias politicas que son VINCULADOS a
    #la facultad de Derecho y Ciencias políticas.

    wb_profe_vincu = load_workbook(profes_vincu)
    sheet_profe_vincu = wb_profe_vincu["reporte"]
    pos_profe_vincu = 'd2'
    pos_num_profe_vincu = 2
    lista_pares_profe_vincu = []
    par_ced_nombre_doc_profe_vincu = []
    lista_pares_nom_ced_profe_vincu = []
    nombre_profe_vincu = []
    while (sheet_profe_vincu[pos_profe_vincu].value != None):
        pos_nom_profe_vincu = 'e' + str(pos_num_profe_vincu)
        nombre_profe_vincu.append(sheet_profe_vincu[pos_nom_profe_vincu].value)
        par_ced_nombre_doc_profe_vincu = []
        for ced_dere in cedula_cp:
            c = sheet_profe_vincu[pos_profe_vincu].value
            if (sheet_profe_vincu[pos_profe_vincu].value == ced_dere):
                par_ced_nombre_doc_profe_vincu.append(ced_dere)
                pos_nom_profe_vincu = 'e' + str(pos_num_profe_vincu)
                par_ced_nombre_doc_profe_vincu.append(sheet_profe_vincu[pos_nom_profe_vincu].value)
                ##pos_doc = 'e' + str(pos_num)
                # par_ced_nombre_doc.append("CC")
                break
        lista_pares_nom_ced_profe_vincu.append(par_ced_nombre_doc_profe_vincu)
        pos_num_profe_vincu = pos_num_profe_vincu + 1
        pos_profe_vincu = 'd' + str(pos_num_profe_vincu)

    nombre_profe_vincu = set(nombre_profe_vincu)

    # Procedimiento para eliminar de lista_pares_nom_ced_profe_vincu las listas que sean vacías.
    resul = []
    for i in lista_pares_nom_ced_profe_vincu:
        if i not in resul:
            resul.append(i)
    res = []
    for i in resul:
        if (i != []):
            res.append(i)

    lista_pares_nom_ced_profe_vincu = res

    #Procedimiento para obtener los profesores de ciencias politicas que son de CATEDRA en
    #la facultad de Derecho y Ciencias Políticas.

    wb = load_workbook(profes_cat)
    sheet = wb["reporte"]
    pos='d2'
    pos_num=2
    lista_pares=[]
    par_ced_nombre_doc=[]
    lista_pares_nom_ced=[]
    nombre=[]
    while(sheet[pos].value!=None):
        pos_nom = 'e' + str(pos_num)
        nombre.append(sheet[pos_nom].value)
        par_ced_nombre_doc = []
        for ced_dere in cedula_cp:
            if(sheet[pos].value==ced_dere):
                par_ced_nombre_doc.append(ced_dere)
                pos_nom='e'+str(pos_num)
                par_ced_nombre_doc.append(sheet[pos_nom].value)
                ##pos_doc = 'e' + str(pos_num)
                #par_ced_nombre_doc.append("CC")
                break
        lista_pares_nom_ced.append(par_ced_nombre_doc)
        pos_num=pos_num+1
        pos='d'+str(pos_num)

    nombre=set(nombre)

    # Procedimiento para eliminar de lista_pares_nom_ced las listas que sean vacías.
    resul = []
    for i in lista_pares_nom_ced:
        if i not in resul:
            resul.append(i)
    res=[]
    for i in resul:
        if(i!=[]):
            res.append(i)

    lista_pares_nom_ced=res

    #Procedimiento para sacar de la lista de CATEDRA a los profesores que son VINCULADOS
    catedra_no_vincu_no_ins=[]
    for i in lista_pares_nom_ced:
        contador_dife=0
        for j in lista_pares_nom_ced_profe_vincu:
            if(j[0]!=i[0]):
                contador_dife+=1
        if(contador_dife==len(lista_pares_nom_ced_profe_vincu)):
            catedra_no_vincu_no_ins.append(i)

    lista_pares_nom_ced=catedra_no_vincu_no_ins

    #Procedimiento para sacar de la lista de CATEDRA a los profesores que son del INSTITUTO
    for i in lista_pares_nom_ced:
        contador_dife = 0
        for j in lista_pares_nom_ced_prof_institu:
            if(j[0]!=i[0]):
                contador_dife+=1
        if(contador_dife==len(lista_pares_nom_ced_prof_institu)):
            pass
        else:
            catedra_no_vincu_no_ins.remove(i)

    lista_pares_nom_ced=catedra_no_vincu_no_ins

    #Procedimiento para sacar de la lista de VINCULADOS los profesores que son del INSTITUTO.
    vincu_no_insti=[]
    for i in lista_pares_nom_ced_profe_vincu:
        contador_dife=0
        for j in lista_pares_nom_ced_prof_institu:
            if(j[0]!=i[0]):
                contador_dife+=1
        if(contador_dife==len(lista_pares_nom_ced_prof_institu)):
            vincu_no_insti.append(i)

    lista_pares_nom_ced_profe_vincu=vincu_no_insti

    res = lista_pares_nom_ced

    catedras_cp=lista_pares_nom_ced
    vinculados_cp=lista_pares_nom_ced_profe_vincu




    #-----------------------------------------------PROFESORES DE DERECHO---------------------------------------------------------------

    #Procedimiento para obtener los nombres de los profesores de derecho que son del instituto de estudios politicos

    wb_prof_institu = load_workbook(prof_institu_cp)
    sheet_prof_institu = wb_prof_institu["reporte"]
    pos_prof_institu = 'd2'
    pos_num_prof_institu = 2
    lista_pares = []
    par_ced_nombre_doc = []
    lista_pares_nom_ced_prof_institu = []
    nombre_prof_institu = []
    while (sheet_prof_institu[pos_prof_institu].value != None):
        pos_nom_prof_institu = 'e' + str(pos_num_prof_institu)
        nombre_prof_institu.append(sheet_prof_institu[pos_nom_prof_institu].value)
        par_ced_nombre_doc_prof_institu = []
        index=0

        for ced_dere in cedula_dere:
            index+=1

            if (sheet_prof_institu[pos_prof_institu].value == ced_dere):
                par_ced_nombre_doc_prof_institu.append(ced_dere)
                pos_nom_prof_institu = 'e' + str(pos_num_prof_institu)
                par_ced_nombre_doc_prof_institu.append(sheet_prof_institu[pos_nom_prof_institu].value)
                break
        lista_pares_nom_ced_prof_institu.append(par_ced_nombre_doc_prof_institu)
        pos_num_prof_institu = pos_num_prof_institu + 1
        pos_prof_institu = 'd' + str(pos_num_prof_institu)

    nombre_prof_institu = set(nombre_prof_institu)

    resul = []
    for i in lista_pares_nom_ced_prof_institu:
        if i not in resul:
            resul.append(i)
    res = []
    for i in resul:
        if (i != []):
            res.append(i)

    lista_pares_nom_ced_prof_institu = res

    #Procedimiento para obtener los nombres de los profesores de derecho que son vinculados

    wb_profe_vincu = load_workbook(profes_vincu_cp)
    sheet_profe_vincu = wb_profe_vincu["reporte"]
    pos_profe_vincu = 'd2'
    pos_num_profe_vincu = 2
    lista_pares_profe_vincu = []
    par_ced_nombre_doc_profe_vincu = []
    lista_pares_nom_ced_profe_vincu = []
    nombre_profe_vincu = []
    while (sheet_profe_vincu[pos_profe_vincu].value != None):
        pos_nom_profe_vincu = 'e' + str(pos_num_profe_vincu)
        nombre_profe_vincu.append(sheet_profe_vincu[pos_nom_profe_vincu].value)
        par_ced_nombre_doc_profe_vincu = []
        for ced_dere in cedula_dere:
            c = sheet_profe_vincu[pos_profe_vincu].value
            if (str(sheet_profe_vincu[pos_profe_vincu].value) == ced_dere):
                par_ced_nombre_doc_profe_vincu.append(ced_dere)
                pos_nom_profe_vincu = 'e' + str(pos_num_profe_vincu)
                par_ced_nombre_doc_profe_vincu.append(sheet_profe_vincu[pos_nom_profe_vincu].value)
                break
        lista_pares_nom_ced_profe_vincu.append(par_ced_nombre_doc_profe_vincu)
        pos_num_profe_vincu = pos_num_profe_vincu + 1
        pos_profe_vincu = 'd' + str(pos_num_profe_vincu)

    nombre_profe_vincu = set(nombre_profe_vincu)

    resul = []
    for i in lista_pares_nom_ced_profe_vincu:
        if i not in resul:
            resul.append(i)
    res = []
    for i in resul:
        if (i != []):
            res.append(i)

    lista_pares_nom_ced_profe_vincu = res

    # Procedimiento para obtener los nombres de los profesores de derecho que son de catedra

    wb = load_workbook(profes_cat_cp)
    sheet = wb["reporte"]
    pos='d2'
    pos_num=2
    lista_pares=[]
    par_ced_nombre_doc=[]
    lista_pares_nom_ced=[]
    nombre=[]
    while(sheet[pos].value!=None):
        pos_nom = 'e' + str(pos_num)
        nombre.append(sheet[pos_nom].value)
        par_ced_nombre_doc = []
        for ced_dere in cedula_dere:
            if(str(sheet[pos].value)==ced_dere):
                par_ced_nombre_doc.append(ced_dere)
                pos_nom='e'+str(pos_num)
                par_ced_nombre_doc.append(sheet[pos_nom].value)
                break
        lista_pares_nom_ced.append(par_ced_nombre_doc)
        pos_num=pos_num+1
        pos='d'+str(pos_num)

    nombre=set(nombre)

    resul = []
    for i in lista_pares_nom_ced:
        if i not in resul:
            resul.append(i)
    res=[]
    for i in resul:
        if(i!=[]):
            res.append(i)

    lista_pares_nom_ced=res

    #Procedimiento para excluir de la lista de profesores de catedra
    #los profesores que son vinculados
    vincu_catedra=[]
    vinculados_dan_catedra=[]
    for i in lista_pares_nom_ced:
        if(i not in lista_pares_nom_ced_profe_vincu):
            vincu_catedra.append(i)
        else:
            vinculados_dan_catedra.append(i)

    lista_pares_nom_ced=vincu_catedra

    res = lista_pares_nom_ced

    #Procedimiento para excluir de la lista de profesores del instituto
    #los profesores que son de catedra o son vinculados
    insti_dere = []
    for i in lista_pares_nom_ced_prof_institu:
        if ((i not in lista_pares_nom_ced_profe_vincu) and (i not in lista_pares_nom_ced)):
            insti_dere.append(i)

    lista_pares_nom_ced_prof_institu=insti_dere

    catedras_derecho=lista_pares_nom_ced
    vinculados_derecho=lista_pares_nom_ced_profe_vincu
    instituto_derecho=lista_pares_nom_ced_prof_institu

    comunes_catedra=[]

    for i in catedras_derecho:
        for j in catedras_cp:
            if(i[0]==j[0]):
                comunes_catedra.append(i)

    comunes_vinculados=[]
    for i in vinculados_derecho:
        for j in vinculados_cp:
            if(i[0]==j[0]):
                comunes_vinculados.append(i)

    for catedras in comunes_catedra:
        for linea in lineas_cedulas:
            if(linea.find(catedras[0])!=-1):
                algovoyahaceraqui=0

    vincu_cpa_horas_cat=[]
    vincu_cpa_horas_plan=[]

    vincu_der_horas_cat = []
    vincu_der_horas_plan = []

    wb_discri = load_workbook(excel_discriminacion)
    sheet_discri= wb_discri["SQL_Results"]

    for ced in comunes_vinculados:
        sum_vincu_cpa_horas_cat = 0
        sum_vincu_cpa_horas_plan = 0
        sum_vincu_der_horas_cat = 0
        sum_vincu_der_horas_plan = 0
        pos_num = 2
        pos = 'A' + str(pos_num)
        while(sheet_discri[pos].value!=None):
            if (str(sheet_discri[pos].value) == str(anio) + str(semest)):
                if (ced[0] == str(sheet_discri["P" + str(pos_num)].value)):
                    if (str(sheet_discri["M" + str(pos_num)].value) == "CPA" or str(sheet_discri["M" + str(pos_num)].value) == "CPT"):
                        if(str(sheet_discri["S"+str(pos_num)].value)!=""):
                            sum_vincu_cpa_horas_cat+=int(sheet_discri["S"+str(pos_num)].value)
                        if(str(sheet_discri["T" + str(pos_num)].value)!=""):
                            sum_vincu_cpa_horas_plan+=int(sheet_discri["T" + str(pos_num)].value)
                    elif (str(sheet_discri["M" + str(pos_num)].value) == "DER" or str(sheet_discri["M" + str(pos_num)].value) == "DEP" or str(sheet_discri["M" + str(pos_num)].value) == "DEI"):
                        if(str(sheet_discri["S"+str(pos_num)].value)!=""):
                            sum_vincu_der_horas_cat+=int(sheet_discri["S"+str(pos_num)].value)
                        if(str(sheet_discri["T" + str(pos_num)].value)!=""):
                            sum_vincu_der_horas_plan+=int(sheet_discri["T" + str(pos_num)].value)
                    else:
                        pass
            pos_num += 1
            pos = 'A' + str(pos_num)
        #Horas de catedra de los profesores vinculados CP
        lista_ind = []
        lista_ind.append(ced[0])
        lista_ind.append(sum_vincu_cpa_horas_cat)
        vincu_cpa_horas_cat.append(lista_ind)
        #Horas del plan vinculados CP
        lista_ind = []
        lista_ind.append(ced[0])
        lista_ind.append(sum_vincu_cpa_horas_plan)
        vincu_cpa_horas_plan.append(lista_ind)
        #Horas de catedra de los profesores vinculados DERECHO
        lista_ind = []
        lista_ind.append(ced[0])
        lista_ind.append(sum_vincu_der_horas_cat)
        vincu_der_horas_cat.append(lista_ind)
        #Horas del plan vinculados DERECHO
        lista_ind = []
        lista_ind.append(ced[0])
        lista_ind.append(sum_vincu_der_horas_plan)
        vincu_der_horas_plan.append(lista_ind)

    cate_cpa_horas_cat = []
    cate_cpa_horas_plan = []
    cate_der_horas_cat = []
    cate_der_horas_plan = []

    for ced in comunes_catedra:
        sum_cate_cpa_horas_cat = 0
        sum_cate_cpa_horas_plan = 0
        sum_cate_der_horas_cat = 0
        sum_cate_der_horas_plan = 0
        pos_num = 2
        pos = 'A' + str(pos_num)
        while(sheet_discri[pos].value!=None):
            if (str(sheet_discri[pos].value) == str(anio) + str(semest)):
                if (ced[0] == str(sheet_discri["P" + str(pos_num)].value)):
                    if (str(sheet_discri["M" + str(pos_num)].value) == "CPA" or str(sheet_discri["M" + str(pos_num)].value) == "CPT"):
                        if(str(sheet_discri["S"+str(pos_num)].value)!=""):
                            sum_cate_cpa_horas_cat+=int(sheet_discri["S"+str(pos_num)].value)
                        if(str(sheet_discri["T" + str(pos_num)].value)!=""):
                            sum_cate_cpa_horas_plan+=int(sheet_discri["T" + str(pos_num)].value)
                    elif (str(sheet_discri["M" + str(pos_num)].value) == "DER" or str(sheet_discri["M" + str(pos_num)].value) == "DEP" or str(sheet_discri["M" + str(pos_num)].value) == "DEI"):
                        if(str(sheet_discri["S"+str(pos_num)].value)!=""):
                            sum_cate_der_horas_cat+=int(sheet_discri["S"+str(pos_num)].value)
                        if(str(sheet_discri["T" + str(pos_num)].value) != ""):
                            sum_cate_der_horas_plan+=int(sheet_discri["T" + str(pos_num)].value)
                    else:
                        pass
            pos_num += 1
            pos = 'A' + str(pos_num)
        #Horas de catedra de los profesores vinculados CP
        lista_ind = []
        lista_ind.append(ced[0])
        lista_ind.append(sum_cate_cpa_horas_cat)
        cate_cpa_horas_cat.append(lista_ind)
        #Horas del plan vinculados CP
        lista_ind = []
        lista_ind.append(ced[0])
        lista_ind.append(sum_cate_cpa_horas_plan)
        cate_cpa_horas_plan.append(lista_ind)
        #Horas de catedra de los profesores vinculados DERECHO
        lista_ind = []
        lista_ind.append(ced[0])
        lista_ind.append(sum_cate_der_horas_cat)
        cate_der_horas_cat.append(lista_ind)
        #Horas del plan vinculados DERECHO
        lista_ind = []
        lista_ind.append(ced[0])
        lista_ind.append(sum_cate_der_horas_plan)
        cate_der_horas_plan.append(lista_ind)

    if(anio==2018 and semest==2):
        print("ojo")

    return cate_cpa_horas_cat,cate_cpa_horas_plan, vincu_cpa_horas_cat,vincu_cpa_horas_plan,cate_der_horas_cat,cate_der_horas_plan,vincu_der_horas_cat,vincu_der_horas_plan


    '''
    while(sheet_discri[pos].value!=None):
        if(str(sheet_discri[pos].value)==str(anio)+str(semest)):
            for ced in comunes_vinculados:
                if(ced[0]==str(sheet_discri["P"+str(pos_num)].value)):
                    if(str(sheet_discri["M"+str(pos_num)].value)=="CPA" or str(sheet_discri["M"+str(pos_num)].value)=="CPT"):
                        lista_ind=[]
                        lista_ind.append(ced[0])
                        lista_ind.append(str(sheet_discri["S"+str(pos_num)].value))
                        vincu_cpa_horas_cat.append(lista_ind)

                        lista_ind = []
                        lista_ind.append(ced[0])
                        lista_ind.append(str(sheet_discri["T" + str(pos_num)].value))
                        vincu_cpa_horas_plan.append(lista_ind)
                    elif(str(sheet_discri["M"+str(pos_num)].value)=="DER" or str(sheet_discri["M"+str(pos_num)].value)=="DEP" or str(sheet_discri["M"+str(pos_num)].value)=="DEI"):
                        lista_ind = []
                        lista_ind.append(ced[0])
                        lista_ind.append(str(sheet_discri["S" + str(pos_num)].value))
                        vincu_der_horas_cat.append(lista_ind)

                        lista_ind = []
                        lista_ind.append(ced[0])
                        lista_ind.append(str(sheet_discri["T" + str(pos_num)].value))
                        vincu_der_horas_plan.append(lista_ind)
                    else:
                        pass
        pos_num+=1
        pos='A'+str(pos_num)
    '''


    s = 1
    s3 = 2