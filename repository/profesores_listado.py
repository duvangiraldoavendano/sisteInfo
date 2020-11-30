from openpyxl import load_workbook
from datetime import datetime

def catedra(profes_cat,profes_vincu, prof_institu,archivo_resul,cedula_dere,
            tipo_cc,anio,semest,reporte_vinculados,planes_derecho,
            lineas_cedulas,cate_der_horas_cat,cate_der_horas_plan,
            vincu_der_horas_cat,vincu_der_horas_plan,excel_discriminacion,
            cate_cpa_horas_cat,cate_cpa_horas_plan,vincu_cpa_horas_cat,vincu_cpa_horas_plan):

    #Procedimiento para obtener los nombres de los profesores de derecho que son del instituto de estudios politicos

    wb_prof_institu = load_workbook(prof_institu)
    sheet_prof_institu = wb_prof_institu["reporte"]
    pos_prof_institu = 'd2'
    pos_num_prof_institu = 2
    lista_pares = []
    par_ced_nombre_doc = []
    lista_pares_nom_ced_prof_institu = []
    nombre_prof_institu = []
    while (sheet_prof_institu[pos_prof_institu].value != None):
        pos_nom_prof_institu = 'e' + str(pos_num_prof_institu)
        nombre_prof_institu.append(sheet_prof_institu[pos_nom_prof_institu].value)
        par_ced_nombre_doc_prof_institu = []
        index=0
        for ced_dere in cedula_dere:
            index+=1
            if (sheet_prof_institu[pos_prof_institu].value == ced_dere):
                par_ced_nombre_doc_prof_institu.append(ced_dere)
                pos_nom_prof_institu = 'e' + str(pos_num_prof_institu)
                par_ced_nombre_doc_prof_institu.append(sheet_prof_institu[pos_nom_prof_institu].value)
                break
        lista_pares_nom_ced_prof_institu.append(par_ced_nombre_doc_prof_institu)
        pos_num_prof_institu = pos_num_prof_institu + 1
        pos_prof_institu = 'd' + str(pos_num_prof_institu)

    nombre_prof_institu = set(nombre_prof_institu)

    resul = []
    for i in lista_pares_nom_ced_prof_institu:
        if i not in resul:
            resul.append(i)
    res = []
    for i in resul:
        if (i != []):
            res.append(i)

    lista_pares_nom_ced_prof_institu = res

    #Procedimiento para obtener los nombres de los profesores de derecho que son vinculados

    wb_profe_vincu = load_workbook(profes_vincu)
    sheet_profe_vincu = wb_profe_vincu["reporte"]
    pos_profe_vincu = 'd2'
    pos_num_profe_vincu = 2
    lista_pares_profe_vincu = []
    par_ced_nombre_doc_profe_vincu = []
    lista_pares_nom_ced_profe_vincu = []
    nombre_profe_vincu = []
    while (sheet_profe_vincu[pos_profe_vincu].value != None):
        pos_nom_profe_vincu = 'e' + str(pos_num_profe_vincu)
        nombre_profe_vincu.append(sheet_profe_vincu[pos_nom_profe_vincu].value)
        par_ced_nombre_doc_profe_vincu = []
        for ced_dere in cedula_dere:
            if (str(sheet_profe_vincu[pos_profe_vincu].value) == ced_dere):
                par_ced_nombre_doc_profe_vincu.append(ced_dere)
                pos_nom_profe_vincu = 'e' + str(pos_num_profe_vincu)
                par_ced_nombre_doc_profe_vincu.append(sheet_profe_vincu[pos_nom_profe_vincu].value)
                break
        lista_pares_nom_ced_profe_vincu.append(par_ced_nombre_doc_profe_vincu)
        pos_num_profe_vincu = pos_num_profe_vincu + 1
        pos_profe_vincu = 'd' + str(pos_num_profe_vincu)

    nombre_profe_vincu = set(nombre_profe_vincu)

    resul = []
    for i in lista_pares_nom_ced_profe_vincu:
        if i not in resul:
            resul.append(i)
    res = []
    for i in resul:
        if (i != []):
            res.append(i)

    lista_pares_nom_ced_profe_vincu = res

    # Procedimiento para obtener los nombres de los profesores de derecho que son de catedra

    wb = load_workbook(profes_cat)
    sheet = wb["reporte"]
    pos='d2'
    pos_num=2
    lista_pares=[]
    par_ced_nombre_doc=[]
    lista_pares_nom_ced=[]
    nombre=[]
    while(sheet[pos].value!=None):
        pos_nom = 'e' + str(pos_num)
        nombre.append(sheet[pos_nom].value)
        par_ced_nombre_doc = []
        for ced_dere in cedula_dere:
            if(str(sheet[pos].value)==ced_dere):
                par_ced_nombre_doc.append(ced_dere)
                pos_nom='e'+str(pos_num)
                par_ced_nombre_doc.append(sheet[pos_nom].value)
                break
        lista_pares_nom_ced.append(par_ced_nombre_doc)
        pos_num=pos_num+1
        pos='d'+str(pos_num)

    nombre=set(nombre)

    resul = []
    for i in lista_pares_nom_ced:
        if i not in resul:
            resul.append(i)
    res=[]
    for i in resul:
        if(i!=[]):
            res.append(i)

    lista_pares_nom_ced=res

    #Procedimiento para excluir de la lista de profesores de catedra
    #los profesores que son vinculados
    vincu_catedra=[]
    vinculados_dan_catedra=[]
    for i in range(0,len(lista_pares_nom_ced)):
        for j in range(0,len(lista_pares_nom_ced_profe_vincu)):
            vincu=0
            if(lista_pares_nom_ced[i][0]==lista_pares_nom_ced_profe_vincu[j][0]):
                vinculados_dan_catedra.append(lista_pares_nom_ced[i])
                vincu=1
                break
        if(vincu==0):
            vincu_catedra.append(lista_pares_nom_ced[i])

    lista_pares_nom_ced=vincu_catedra
    res = lista_pares_nom_ced

    #Procedimiento para excluir de la lista de profesores del instituto
    #los profesores que son de catedra o son vinculados
    insti_dere = []
    for i in lista_pares_nom_ced_prof_institu:
        if ((i not in lista_pares_nom_ced_profe_vincu) and (i not in lista_pares_nom_ced)):
            insti_dere.append(i)

    lista_pares_nom_ced_prof_institu=insti_dere

    #Procedimiento para obtener el tipo de documento de los profesores de catedra
    wb_tip_doc = load_workbook(tipo_cc)
    sheet_tipdoc = wb_tip_doc["reporte"]
    indice=0

    for profe in lista_pares_nom_ced:
        pos_num = 2
        pos = 'f' + str(pos_num)
        cont_ocur=0
        while (sheet_tipdoc[pos].value != None):
            if (str(sheet_tipdoc[pos].value) == profe[0]):
                lista_pares_nom_ced[indice].append(sheet_tipdoc['e' + str(pos_num)].value)
                cont_ocur+=1
                break
            pos_num += 1
            pos = 'f' + str(pos_num)
        if(cont_ocur==0):
            lista_pares_nom_ced[indice].append(" ")
        indice += 1

    areas_conocimiento=('AGRONOMÍA, VETERINARIA Y AFINES',
                        'BELLAS ARTES',
                        'CIENCIAS DE LA EDUCACIÓN',
                        'CIENCIAS DE LA SALUD',
                        'CIENCIAS SOCIALES, DERECHO, CIENCIAS POLÍTICAS',
                        'ECONOMÍA, ADMINISTRACIÓN, CONTADURÍA Y AFINES',
                        'INGENIERÍA, ARQUITECTURA, URBANISMO Y AFINES',
                        'MATEMÁTICAS Y CIENCIAS NATURALES',
                        'HUMANIDADES Y CIENCIAS RELIGIOSAS')

    #Funcion que recibe el titulo de un profesor y retorna el area de conocimiento en la que se encuentra
    def area_conoci(titulo):
        if (('PSICO' in titulo) or ('SICOL' in titulo) or ('CONTEMP' in titulo) or('CONFLIC' in titulo) or ('HISTO' in titulo) or ('POL' in titulo) or ('DER' in titulo) or ('ABOG' in titulo) or ('SOCI' in titulo) or ('IDEOL' in titulo) or ('FISC' in titulo) or ('JURID' in titulo) or ('GOBI' in titulo) or ('PENA' in titulo) or ('PENIT' in titulo) or ('CIVI' in titulo) or ('DESARRO' in titulo) or ('ESTADO' in titulo) or ('RELACIO' in titulo) or ('LATINOAMER' in titulo) or ('HABIT' in titulo)):
            return areas_conocimiento[4]
        elif (('EDU' in titulo) or ('PEDAG' in titulo) or ('LINGU' in titulo) or ('BIBLIO' in titulo)):
            return areas_conocimiento[2]
        elif (('ADMI' in titulo) or ('CONTAD' in titulo) or ('ECONO' in titulo) or ('GEREN' in titulo) or ('FINAN' in titulo) or ('GESTIO' in titulo)  or ('CONTRAT' in titulo)):
            return areas_conocimiento[5]
        elif (('FAMILI' in titulo)  or ('FILOS' in titulo) or ('HUMAN' in titulo)):
            return areas_conocimiento[8]
        elif (('SALUD' in titulo) or ('MEDIC' in titulo)):
            return areas_conocimiento[3]
        elif (('CIENCI' in titulo)):
            return areas_conocimiento[7]
        elif (('INFOR' in titulo) or ('AMBIE' in titulo) or ('TECNOL' in titulo) or ('URBA' in titulo)):
            return areas_conocimiento[6]
        if ( ('FAMILI' in titulo)  or ('FILOS' in titulo) or ('HUMAN' in titulo)):
            return areas_conocimiento[8]
        elif(('ESCENI' in titulo)):
            return areas_conocimiento[1]
        else:
            return " "

    #Procedimiento para escoger el titulo vigente más alto de los profesores de catedra
    indice = 0
    for ced in lista_pares_nom_ced:
        titulos=[]
        pos = 'd2'
        pos_num = 2
        while(sheet[pos].value!=None):
            if(ced[0]==str(sheet[pos].value)):
                titulos.append(sheet["f"+str(pos_num)].value)
                titulos.append(pos_num)
            pos_num+=1
            pos="d"+str(pos_num)
        if (semest==1):
            mes=7
        else:
            mes=12
        if "DOCTO" in titulos:
            if((int(str(sheet["g" + str(titulos[titulos.index("DOCTO") + 1])].value)[0:4]))<anio):
                vigen_docto=True
            elif((int(str(sheet["g" + str(titulos[titulos.index("DOCTO") + 1])].value)[0:4]))==anio):
                if((int(str(sheet["g" + str(titulos[titulos.index("DOCTO") + 1])].value)[5:7]))<=mes):
                    vigen_docto=True
                else:
                    vigen_docto=False
            else:
                vigen_docto=False
        if("MAEST" in titulos):
            if ((int(str(sheet["g" + str(titulos[titulos.index("MAEST") + 1])].value)[0:4])) < anio):
                vigen_maest = True
            elif ((int(str(sheet["g" + str(titulos[titulos.index("MAEST") + 1])].value)[0:4])) == anio):
                if ((int(str(sheet["g" + str(titulos[titulos.index("MAEST") + 1])].value)[5:7])) <= mes):
                    vigen_maest = True
                else:
                    vigen_maest = False
            else:
                vigen_maest = False
        if("ESPEC" in titulos):
            if ((int(str(sheet["g" + str(titulos[titulos.index("ESPEC") + 1])].value)[0:4])) < anio):
                vigen_espe = True
            elif ((int(str(sheet["g" + str(titulos[titulos.index("ESPEC") + 1])].value)[0:4])) == anio):
                if ((int(str(sheet["g" + str(titulos[titulos.index("ESPEC") + 1])].value)[5:7])) <= mes):
                    vigen_espe = True
                else:
                    vigen_espe = False
            else:
                vigen_espe = False
        if("PREGR" in titulos):
            if ((int(str(sheet["g" + str(titulos[titulos.index("PREGR") + 1])].value)[0:4])) < anio):
                vigen_pregr = True
            elif ((int(str(sheet["g" + str(titulos[titulos.index("PREGR") + 1])].value)[0:4])) == anio):
                if ((int(str(sheet["g" + str(titulos[titulos.index("PREGR") + 1])].value)[5:7])) <= mes):
                    vigen_pregr = True
                else:
                    vigen_pregr = False
            else:
                vigen_pregr = False
        if(("DOCTO" in titulos) and vigen_docto):
            lista_pares_nom_ced[indice].append("DOCTORADO")
            titulo = sheet["i" + str(titulos[titulos.index("DOCTO") + 1])].value
            if (titulo.find("DOCTORADO") != -1):
                lista_pares_nom_ced[indice].append("DOCTOR(A) " + titulo[10:len(titulo)])
            elif (titulo.find("DOCTOR(A)") != -1):
                lista_pares_nom_ced[indice].append("DOCTOR(A) " + titulo[10:len(titulo)])
            elif (titulo.find("DOCTOR (A)") != -1):
                lista_pares_nom_ced[indice].append("DOCTOR(A) " + titulo[11:len(titulo)])
            elif (titulo.find("DOCTOR") != -1):
                lista_pares_nom_ced[indice].append("DOCTOR(A) " + titulo[7:len(titulo)])
            elif (titulo.find("DOCTORA") != -1):
                lista_pares_nom_ced[indice].append("DOCTOR(A) " + titulo[8:len(titulo)])
            else:
                lista_pares_nom_ced[indice].append(titulo)

            separacion = (sheet["j" + str(titulos[titulos.index("DOCTO") + 1])].value)
            if ((separacion.find("UNIVERSIDAD")!=-1) and (separacion.find("EAFIT")!=-1)):
                lista_pares_nom_ced[indice].append((sheet["j" + str(titulos[titulos.index("DOCTO") + 1])].value)[0:17])
            elif ((separacion.find("UNIVERSIDAD")!=-1) and (separacion.find("NACIONAL")!=-1) and (separacion.find("COLOMBIA")!=-1)):
                lista_pares_nom_ced[indice].append((sheet["j" + str(titulos[titulos.index("DOCTO") + 1])].value)[0:32])
            else:
                lista_pares_nom_ced[indice].append((sheet["j" + str(titulos[titulos.index("DOCTO") + 1])].value))

            lista_pares_nom_ced[indice].append(area_conoci(titulo))

        elif(("MAEST" in titulos) and vigen_maest):
            lista_pares_nom_ced[indice].append("MAESTRÍA")
            titulo = sheet["i" + str(titulos[titulos.index("MAEST") + 1])].value
            if (titulo.find("MAESTRIA") != -1 ):
                lista_pares_nom_ced[indice].append("MAGÍSTER" + titulo[8:len(titulo)])
            elif(titulo.find("MASTER") != -1 ):
                lista_pares_nom_ced[indice].append("MAGÍSTER" + titulo[6:len(titulo)])
            else:
                lista_pares_nom_ced[indice].append(titulo)

            separacion=(sheet["j" + str(titulos[titulos.index("MAEST") + 1])].value)
            if ((separacion.find("UNIVERSIDAD")!=-1) and (separacion.find("EAFIT")!=-1)):
                lista_pares_nom_ced[indice].append((sheet["j" + str(titulos[titulos.index("MAEST") + 1])].value)[0:17])
            elif ((separacion.find("UNIVERSIDAD") != -1) and (separacion.find("NACIONAL") != -1) and (separacion.find("COLOMBIA") != -1)):
                lista_pares_nom_ced[indice].append((sheet["j" + str(titulos[titulos.index("MAEST") + 1])].value)[0:32])
            else:
                lista_pares_nom_ced[indice].append((sheet["j" + str(titulos[titulos.index("MAEST") + 1])].value))

            lista_pares_nom_ced[indice].append(area_conoci(titulo))

        elif (("ESPEC" in titulos) and vigen_espe):
            lista_pares_nom_ced[indice].append("ESPECIALIZACIÓN")
            titulo = sheet["i" + str(titulos[titulos.index("ESPEC") + 1])].value
            if(titulo.find("ESP.")!=-1):
                lista_pares_nom_ced[indice].append("ESPECIALISTA"+titulo[4:len(titulo)])
            elif(titulo.find("ESPECIALI")!=-1):
                lista_pares_nom_ced[indice].append("ESPECIALISTA" + titulo[15:len(titulo)])
            else:
                lista_pares_nom_ced[indice].append(titulo)

            separacion = (sheet["j" + str(titulos[titulos.index("ESPEC") + 1])].value)
            if ((separacion.find("UNIVERSIDAD")!=-1) and (separacion.find("EAFIT")!=-1)):
                lista_pares_nom_ced[indice].append((sheet["j" + str(titulos[titulos.index("ESPEC") + 1])].value)[0:17])
            elif ((separacion.find("UNIVERSIDAD") != -1) and (separacion.find("NACIONAL") != -1) and (
                            separacion.find("COLOMBIA") != -1)):
                lista_pares_nom_ced[indice].append((sheet["j" + str(titulos[titulos.index("ESPEC") + 1])].value)[0:32])
            else:
                lista_pares_nom_ced[indice].append(sheet["j" + str(titulos[titulos.index("ESPEC") + 1])].value)

            lista_pares_nom_ced[indice].append(area_conoci(titulo))

        elif(("PREGR" in titulos) and vigen_pregr):
            lista_pares_nom_ced[indice].append("PREGRADO")
            if(sheet["i"+str(titulos[titulos.index("PREGR")+1])].value == "DERECHO"):
                lista_pares_nom_ced[indice].append("ABOGADO")
            else:
                lista_pares_nom_ced[indice].append(sheet["i"+str(titulos[titulos.index("PREGR")+1])].value)

            separacion = (sheet["j" + str(titulos[titulos.index("PREGR") + 1])].value)
            if ((separacion.find("UNIVERSIDAD")!=-1) and (separacion.find("EAFIT")!=-1)):
                lista_pares_nom_ced[indice].append((sheet["j" + str(titulos[titulos.index("PREGR") + 1])].value)[0:17])
            elif ((separacion.find("UNIVERSIDAD") != -1) and (separacion.find("NACIONAL") != -1) and (
                            separacion.find("COLOMBIA") != -1)):
                lista_pares_nom_ced[indice].append((sheet["j" + str(titulos[titulos.index("PREGR") + 1])].value)[0:32])
            else:
                lista_pares_nom_ced[indice].append((sheet["j" + str(titulos[titulos.index("PREGR") + 1])].value))

            titulo = sheet["i" + str(titulos[titulos.index("PREGR") + 1])].value
            lista_pares_nom_ced[indice].append(area_conoci(titulo))
        else:
            pass
        indice+=1

    #Procedimiento para separar el nombre del apellido de los profesores
    apellidos=[]
    nombres=[]
    for f in range(len(res)):
        noms_apelli=res[f][1].split(" ")
        if(("CASTI" in noms_apelli) and ("LONGAS" in noms_apelli) and ("ANDRES" in noms_apelli)):
            apellidos.append("DIAZ DEL CASTILLO LONGAS")
            nombres.append("ANDRES")
        elif(("D" in noms_apelli) and ("DUQUE" in noms_apelli) and("CAMILO" in noms_apelli)):
            apellidos.append("ARANGO DUQUE")
            nombres.append("CAMILO")
        elif(len(noms_apelli)==3):
            apellidos.append(noms_apelli[0]+" "+noms_apelli[1])
            nombres.append(noms_apelli[2])
        elif(len(noms_apelli)==2):
            apellidos.append(noms_apelli[0])
            nombres.append(noms_apelli[1])
        elif(len(noms_apelli)==4):
            if(noms_apelli[1]=="DE" or noms_apelli[1]=="de"):
                apellidos.append(noms_apelli[0] + " " + noms_apelli[1]+" "+noms_apelli[2])
                nombres.append(noms_apelli[3])
            else:
                apellidos.append(noms_apelli[0] + " " + noms_apelli[1])
                nombres.append(noms_apelli[2]+" "+noms_apelli[3])
        elif(len(noms_apelli)==1):
            nombres.append(noms_apelli[0])
        elif(len(noms_apelli)>=5):
            nombres_p = ""
            if (noms_apelli[1] == "DE" or noms_apelli[1] == "de"):
                apellidos.append(noms_apelli[0] + " " + noms_apelli[1] + " " + noms_apelli[2])
                for i in range(3, len(noms_apelli)):
                    nombres_p = nombres_p + noms_apelli[i] + " "
                nombres.append(nombres_p)
            else:
                apellidos.append(noms_apelli[0]+" "+noms_apelli[1])
                nombres_p = ""
                for i in range(2,len(noms_apelli)):
                    nombres_p=nombres_p+noms_apelli[i]+" "
                nombres.append(nombres_p)
        else:
            pass

    #Procedimiento para encontrar las horas de docencia de los profesores de catedra

    wb_tip_doc = load_workbook(tipo_cc)
    sheet_tipdoc = wb_tip_doc["reporte"]
    indice = 0
    par_ced_horas_doce_catedra=[]
    par_ced_horas_ext_catedra=[]
    par_ced_horas_inv_catedra=[]
    for profe in lista_pares_nom_ced:
        pos_num = 2
        pos = 'f' + str(pos_num)
        horas_docencia_catedra=0
        horas_extension_catedra = 0#Hasta el 20182 los profesores de catedra solo podían hacer docencia y extesion
        while (sheet_tipdoc[pos].value != None):
            if (str(sheet_tipdoc[pos].value) == profe[0]):
                if((str(sheet_tipdoc['c'+str(pos_num)].value) =='21820001') or (str(sheet_tipdoc['c'+str(pos_num)].value) =='21820002') or(str(sheet_tipdoc['c'+str(pos_num)].value) =='21840001')):
                    if(str(sheet_tipdoc['l'+str(pos_num)].value)=='PREGR'):
                        horas_docencia_catedra+=int(str(sheet_tipdoc['m'+str(pos_num)].value))
                elif( (str(sheet_tipdoc['c'+str(pos_num)].value) =='21860002') ):
                    if (str(sheet_tipdoc['l' + str(pos_num)].value) == 'PREGR' or str(sheet_tipdoc['l' + str(pos_num)].value) == 'EXTEN'):
                        horas_extension_catedra+=int(str(sheet_tipdoc['m'+str(pos_num)].value))
                else:
                    pass
            pos_num += 1
            pos = 'f' + str(pos_num)
        par0=[]
        par0.append(profe[0])
        par0.append(horas_docencia_catedra)
        par_ced_horas_doce_catedra.append(par0)

        par_ext=[]
        par_ext.append(profe[0])
        par_ext.append(horas_extension_catedra)
        par_ced_horas_ext_catedra.append(par_ext)

        if(anio<=2018):
            par_inv = []
            par_inv.append(profe[0])
            par_inv.append(0)
            par_ced_horas_inv_catedra.append(par_inv)
        indice += 1

    profes_vige_futuro=[]
    for i in res:
        if(len(i)==7):
            profes_vige_futuro.append(i)
        elif(len(i)==3):
            p=i
            p.append("No registra")
            p.append("No registra")
            p.append("No registra")
            p.append("No registra")
            profes_vige_futuro.append(p)
        else:
            pass

    res=profes_vige_futuro

    #Procedimiento para discriminar las horas de aquellos profesores que tienen horas
    #en el pregrado de derecho y el de Ciencias políticas.

    docencia_catedra=par_ced_horas_doce_catedra
    for i in range(0,len(docencia_catedra)):
        for j in range(0,len(cate_der_horas_cat)):
            if(docencia_catedra[i][0]==cate_der_horas_cat[j][0]):
                resta = par_ced_horas_doce_catedra[i][1] - (cate_cpa_horas_cat[j][1] + cate_cpa_horas_plan[j][1])
                if(cate_der_horas_cat[j][1]==0 and cate_der_horas_plan[j][1]==0 and resta>0):
                    par_ced_horas_doce_catedra[i][1] = resta
                else:
                    par_ced_horas_doce_catedra[i][1]=cate_der_horas_cat[j][1]+cate_der_horas_plan[j][1]
                break

    wb_discri = load_workbook(excel_discriminacion)
    sheet_discri = wb_discri["SQL_Results"]

    copia_docencia = par_ced_horas_doce_catedra
    indice = 0
    for ced in copia_docencia:
        if (ced[1] == None or (ced[1] == 0 and par_ced_horas_ext_catedra[indice][1] == 0)):
            sum_vincu_der_horas_cat = 0
            sum_vincu_der_horas_plan = 0
            pos_num = 2
            pos = 'A' + str(pos_num)
            while (sheet_discri[pos].value != None):
                if (str(sheet_discri[pos].value) == str(anio) + str(semest)):
                    if (ced[0] == str(sheet_discri["P" + str(pos_num)].value)):
                        if (str(sheet_discri["M" + str(pos_num)].value) == "DER" or str(
                                sheet_discri["M" + str(pos_num)].value) == "DEP" or str(
                            sheet_discri["M" + str(pos_num)].value) == "DEI"):
                            if (str(sheet_discri["S" + str(pos_num)].value) != ""):
                                sum_vincu_der_horas_cat += int(sheet_discri["S" + str(pos_num)].value)
                            if (str(sheet_discri["T" + str(pos_num)].value) != ""):
                                sum_vincu_der_horas_plan += int(sheet_discri["T" + str(pos_num)].value)
                pos_num += 1
                pos = 'A' + str(pos_num)
            # Horas de catedra de los profesores de catedra DERECHO
            par_ced_horas_doce_catedra[indice][1] = sum_vincu_der_horas_cat
        indice += 1

    #Procedimiento para guardar en el archivo de resultado, la informacion de los profesores de catedra
    #nombre, apellido, tipo de documento, nivel academico, titulo obtenido, area del conocimiento e institucion donde realizo los estudios
    wb = load_workbook(archivo_resul)
    sheet = wb["profesores"]
    for i in range(len(res)):

        z = 'a' + str(i + 2)
        if(i+1==222):
            print("Hi")
        sheet[z] = i+1

        z='i'+ str(i + 2)
        sheet[z]=res[i][5]

        z = 'g' + str(i + 2)
        sheet[z] = res[i][6]

        z = 'h' + str(i + 2)
        sheet[z] = res[i][4]

        z='e'+str(i+2)
        sheet[z]=res[i][0]

        h = 'd' + str(i + 2)
        sheet[h] = res[i][2]

        x='c'+str(i+2)
        sheet[x]=apellidos[i]

        y='b'+str(i+2)
        sheet[y] = nombres[i]

        y = 'f' + str(i + 2)
        sheet[y] = res[i][3]

        y="j"+str(i+2)
        sheet[y]="CATEDRA"

        y = "o" + str(i + 2)
        sheet[y] = "TERMINO FIJO"

        y = "p" + str(i + 2)
        sheet[y] = "MESES"

        y = "q" + str(i + 2)
        sheet[y] = 5

        y='l'+str(i+2)
        sheet[y] = par_ced_horas_doce_catedra[i][1]

        y = 'n' + str(i + 2)
        sheet[y] = par_ced_horas_ext_catedra[i][1]

        if(anio<=2018):
            y = 'm' + str(i + 2)
            sheet[y] = par_ced_horas_inv_catedra[i][1]

            y = 'k' + str(i + 2)
            sheet[y] = par_ced_horas_doce_catedra[i][1] + par_ced_horas_ext_catedra[i][1] + par_ced_horas_inv_catedra[i][1]

    wb.save(archivo_resul)

    #Procedimiento para obtener el tipo de documento de los profesores vinculados,tambien para saber la fecha
    #de vinculacion, y la clase de empleado que es

    wb_tip_doc = load_workbook(reporte_vinculados)
    sheet_tipdoc = wb_tip_doc["reporte"]
    indice=0
    dias_vinculacion=[]
    clase_empleado=[]
    mes_anio=[]
    for profe in lista_pares_nom_ced_profe_vincu:
        pos_num = 2
        pos = 'f' + str(pos_num)
        while (sheet_tipdoc[pos].value != None):
            if (str(sheet_tipdoc[pos].value) == profe[0]):
                fecha_actual = datetime(anio, mes, 30)
                fecha_excel=sheet_tipdoc['k' + str(pos_num)].value
                resta=str((fecha_actual-fecha_excel))
                resta_fechas=''
                for i in resta:
                    if (i!='d'):
                        resta_fechas=resta_fechas+i
                    else:
                        break
                resta_fechas=int(resta_fechas)

                if(resta_fechas<365):
                    mes_anio.append(0)
                    dias_vinculacion.append(int(resta_fechas/30))
                else:
                    mes_anio.append(1)
                    dias_vinculacion.append(int(resta_fechas/365))

                clase_empleado.append(sheet_tipdoc['i' + str(pos_num)].value)
                break
            pos_num += 1
            pos = 'f' + str(pos_num)
        indice += 1

    indice = 0
    wb_tip_doc = load_workbook(reporte_vinculados)
    sheet_tipdoc = wb_tip_doc["reporte"]
    indice = 0
    for profe in lista_pares_nom_ced_profe_vincu:
        pos_num = 2
        pos = 'f' + str(pos_num)
        while (sheet_tipdoc[pos].value != None):
            if (sheet_tipdoc[pos].value == profe[0]):
                lista_pares_nom_ced_profe_vincu[indice].append(sheet_tipdoc['e' + str(pos_num)].value)
                break
            pos_num += 1
            pos = 'f' + str(pos_num)
        indice += 1

    indice = 0
    #Procedimiento para escoger el titulo vigente mas alto para profesores vinculados
    wb_profe_vincu = load_workbook(profes_vincu)
    sheet = wb_profe_vincu["reporte"]
    res = lista_pares_nom_ced_profe_vincu
    for ced in lista_pares_nom_ced_profe_vincu:
        titulos = []
        pos = 'd2'
        pos_num = 2
        while (sheet[pos].value != None):
            if (ced[0] == str(sheet[pos].value)):
                titulos.append(sheet["f" + str(pos_num)].value)
                titulos.append(pos_num)
            pos_num += 1
            pos = "d" + str(pos_num)
        if (semest == 1):
            mes = 7
        else:
            mes = 12
        if "DOCTO" in titulos:
            if ((int(str(sheet["g" + str(titulos[titulos.index("DOCTO") + 1])].value)[0:4])) < anio):
                vigen_docto = True
            elif ((int(str(sheet["g" + str(titulos[titulos.index("DOCTO") + 1])].value)[0:4])) == anio):
                if ((int(str(sheet["g" + str(titulos[titulos.index("DOCTO") + 1])].value)[5:7])) <= mes):
                    vigen_docto = True
                else:
                    vigen_docto = False
            else:
                vigen_docto = False
        if ("MAEST" in titulos):
            if ((int(str(sheet["g" + str(titulos[titulos.index("MAEST") + 1])].value)[0:4])) < anio):
                vigen_maest = True
            elif ((int(str(sheet["g" + str(titulos[titulos.index("MAEST") + 1])].value)[0:4])) == anio):
                if ((int(str(sheet["g" + str(titulos[titulos.index("MAEST") + 1])].value)[5:7])) <= mes):
                    vigen_maest = True
                else:
                    vigen_maest = False
            else:
                vigen_maest = False
        if ("ESPEC" in titulos):
            if ((int(str(sheet["g" + str(titulos[titulos.index("ESPEC") + 1])].value)[0:4])) < anio):
                vigen_espe = True
            elif ((int(str(sheet["g" + str(titulos[titulos.index("ESPEC") + 1])].value)[0:4])) == anio):
                if ((int(str(sheet["g" + str(titulos[titulos.index("ESPEC") + 1])].value)[5:7])) <= mes):
                    vigen_espe = True
                else:
                    vigen_espe = False
            else:
                vigen_espe = False
        if ("PREGR" in titulos):
            if ((int(str(sheet["g" + str(titulos[titulos.index("PREGR") + 1])].value)[0:4])) < anio):
                vigen_pregr = True
            elif ((int(str(sheet["g" + str(titulos[titulos.index("PREGR") + 1])].value)[0:4])) == anio):
                if ((int(str(sheet["g" + str(titulos[titulos.index("PREGR") + 1])].value)[5:7])) <= mes):
                    vigen_pregr = True
                else:
                    vigen_pregr = False
            else:
                vigen_pregr = False
        if (("DOCTO" in titulos) and vigen_docto):
            lista_pares_nom_ced_profe_vincu[indice].append("DOCTORADO")
            titulo = sheet["i" + str(titulos[titulos.index("DOCTO") + 1])].value
            if (titulo.find("DOCTORADO") != -1):
                lista_pares_nom_ced_profe_vincu[indice].append("DOCTOR(A) " + titulo[10:len(titulo)])
            elif (titulo.find("DOCTOR(A)") != -1):
                lista_pares_nom_ced_profe_vincu[indice].append("DOCTOR(A) " + titulo[10:len(titulo)])
            elif (titulo.find("DOCTOR (A)") != -1):
                lista_pares_nom_ced_profe_vincu[indice].append("DOCTOR(A) " + titulo[11:len(titulo)])
            elif (titulo.find("DOCTOR") != -1):
                lista_pares_nom_ced_profe_vincu[indice].append("DOCTOR(A) " + titulo[7:len(titulo)])
            elif (titulo.find("DOCTORA") != -1):
                lista_pares_nom_ced_profe_vincu[indice].append("DOCTOR(A) " + titulo[8:len(titulo)])
            else:
                lista_pares_nom_ced_profe_vincu[indice].append(titulo)

            separacion = (sheet["j" + str(titulos[titulos.index("DOCTO") + 1])].value)
            if ((separacion.find("UNIVERSIDAD")!=-1) and (separacion.find("EAFIT")!=-1)):
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("DOCTO") + 1])].value)[0:17])
            elif ((separacion.find("UNIVERSIDAD") != -1) and (separacion.find("NACIONAL") != -1) and (
                            separacion.find("COLOMBIA") != -1)):
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("DOCTO") + 1])].value)[0:32])
            else:
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("DOCTO") + 1])].value))

            lista_pares_nom_ced_profe_vincu[indice].append(area_conoci(titulo))

        elif (("MAEST" in titulos) and vigen_maest):
            lista_pares_nom_ced_profe_vincu[indice].append("MAESTRÍA")
            titulo = sheet["i" + str(titulos[titulos.index("MAEST") + 1])].value
            if (titulo.find("MAESTRIA") != -1):
                lista_pares_nom_ced_profe_vincu[indice].append("MAGÍSTER" + titulo[8:len(titulo)])
            elif (titulo.find("MASTER") != -1):
                lista_pares_nom_ced_profe_vincu[indice].append("MAGÍSTER" + titulo[6:len(titulo)])
            else:
                lista_pares_nom_ced_profe_vincu[indice].append(titulo)

            separacion = (sheet["j" + str(titulos[titulos.index("MAEST") + 1])].value)
            if ((separacion.find("UNIVERSIDAD")!=-1) and (separacion.find("EAFIT")!=-1)):
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("MAEST") + 1])].value)[0:17])
            elif ((separacion.find("UNIVERSIDAD") != -1) and (separacion.find("NACIONAL") != -1) and (
                            separacion.find("COLOMBIA") != -1)):
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("MAEST") + 1])].value)[0:32])
            else:
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("MAEST") + 1])].value))

            lista_pares_nom_ced_profe_vincu[indice].append(area_conoci(titulo))

        elif (("ESPEC" in titulos) and vigen_espe):
            lista_pares_nom_ced_profe_vincu[indice].append("ESPECIALIZACIÓN")
            titulo = sheet["i" + str(titulos[titulos.index("ESPEC") + 1])].value
            if (titulo.find("ESP.") != -1):
                lista_pares_nom_ced_profe_vincu[indice].append("ESPECIALISTA" + titulo[4:len(titulo)])
            elif (titulo.find("ESPECIALI") != -1):
                lista_pares_nom_ced_profe_vincu[indice].append("ESPECIALISTA" + titulo[15:len(titulo)])
            else:
                lista_pares_nom_ced_profe_vincu[indice].append(titulo)

            separacion = (sheet["j" + str(titulos[titulos.index("ESPEC") + 1])].value)
            if ((separacion.find("UNIVERSIDAD")!=-1) and (separacion.find("EAFIT")!=-1)):
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("ESPEC") + 1])].value)[0:17])
            elif ((separacion.find("UNIVERSIDAD") != -1) and (separacion.find("NACIONAL") != -1) and (
                            separacion.find("COLOMBIA") != -1)):
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("ESPEC") + 1])].value)[0:32])
            else:
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("ESPEC") + 1])].value))

            lista_pares_nom_ced_profe_vincu[indice].append(area_conoci(titulo))

        elif (("PREGR" in titulos) and vigen_pregr):
            lista_pares_nom_ced_profe_vincu[indice].append("PREGRADO")
            if (sheet["i" + str(titulos[titulos.index("PREGR") + 1])].value == "DERECHO"):
                lista_pares_nom_ced_profe_vincu[indice].append("ABOGADO")
            else:
                lista_pares_nom_ced_profe_vincu[indice].append(sheet["i" + str(titulos[titulos.index("PREGR") + 1])].value)

            separacion = (sheet["j" + str(titulos[titulos.index("PREGR") + 1])].value)
            if ((separacion.find("UNIVERSIDAD")!=-1) and (separacion.find("EAFIT")!=-1)):
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("PREGR") + 1])].value)[0:17])
            elif ((separacion.find("UNIVERSIDAD") != -1) and (separacion.find("NACIONAL") != -1) and (
                            separacion.find("COLOMBIA") != -1)):
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("PREGR") + 1])].value)[0:32])
            else:
                lista_pares_nom_ced_profe_vincu[indice].append((sheet["j" + str(titulos[titulos.index("PREGR") + 1])].value))

            titulo = sheet["i" + str(titulos[titulos.index("PREGR") + 1])].value
            lista_pares_nom_ced_profe_vincu[indice].append(area_conoci(titulo))
        else:
            pass
        indice += 1

    res = lista_pares_nom_ced_profe_vincu

    #Procedimiento para separar nombres de apellidos para los profesores vinculados
    apellidos = []
    nombres = []
    for f in range(len(res)):
        noms_apelli = res[f][1].split(" ")
        if (("CASTI" in noms_apelli) and ("LONGAS" in noms_apelli) and ("ANDRES" in noms_apelli)):
            apellidos.append("DIAZ DEL CASTILLO LONGAS")
            nombres.append("ANDRES")
        elif (("D" in noms_apelli) and ("DUQUE" in noms_apelli) and ("CAMILO" in noms_apelli)):
            apellidos.append("ARANGO DUQUE")
            nombres.append("CAMILO")
        elif (len(noms_apelli) == 3):
            apellidos.append(noms_apelli[0] + " " + noms_apelli[1])
            nombres.append(noms_apelli[2])
        elif (len(noms_apelli) == 2):
            apellidos.append(noms_apelli[0])
            nombres.append(noms_apelli[1])
        elif (len(noms_apelli) == 4):
            if (noms_apelli[1] == "DE" or noms_apelli[1] == "de"):
                apellidos.append(noms_apelli[0] + " " + noms_apelli[1] + " " + noms_apelli[2])
                nombres.append(noms_apelli[3])
            else:
                apellidos.append(noms_apelli[0] + " " + noms_apelli[1])
                nombres.append(noms_apelli[2] + " " + noms_apelli[3])
        elif (len(noms_apelli) == 1):
            nombres.append(noms_apelli[0])
        elif (len(noms_apelli) >= 5):
            nombres_p = ""
            if (noms_apelli[1] == "DE" or noms_apelli[1] == "de"):
                apellidos.append(noms_apelli[0] + " " + noms_apelli[1] + " " + noms_apelli[2])
                for i in range(3, len(noms_apelli)):
                    nombres_p = nombres_p + noms_apelli[i] + " "
                nombres.append(nombres_p)
            else:
                apellidos.append(noms_apelli[0] + " " + noms_apelli[1])
                nombres_p = ""
                for i in range(2, len(noms_apelli)):
                    nombres_p = nombres_p + noms_apelli[i] + " "
                nombres.append(nombres_p)
        else:
            pass

    #Procedimiento para obtener la dedicacion de los profesores vinculados

    wb_dedica_vincu = load_workbook(reporte_vinculados)
    sheet_dedica_vincu = wb_dedica_vincu["reporte"]
    indice = 0
    dedicacion = []
    for profe in lista_pares_nom_ced_profe_vincu:
        pos_num = 2
        pos = 'f' + str(pos_num)
        while (sheet_dedica_vincu[pos].value != None):
            if (str(sheet_dedica_vincu[pos].value) == profe[0]):
                dedicacion.append(sheet_dedica_vincu['j' + str(pos_num)].value)
                break
            pos_num += 1
            pos = 'f' + str(pos_num)
        indice += 1


    #Procedimiento para obtener las horas dedicadas a docencia, extension e investigacion de los profesores vinculados

    if anio<=2017 and semest<=1:
        doc='n'
        inv='p'
        ext='q'
        docum='c'
        hoja='plan_x_prof'
    else:
        doc = 'f'
        inv = 'h'
        ext = 'j'
        docum= 'b'
        hoja='reporte_resumido'

    wb_horas = load_workbook(planes_derecho)
    sheet_horas = wb_horas[hoja]
    docencia = []
    investigacion = []
    extension = []
    indice=0

    for profe in lista_pares_nom_ced_profe_vincu:
        pos_num = 2
        pos = docum + str(pos_num)
        no_aparece=1
        doc_acum=0
        inv_acum=0
        ext_acum=0
        while (sheet_horas[pos].value != None):
            par1=[]
            par2=[]
            par3=[]

            if (str(sheet_horas[pos].value) == profe[0]):
                no_aparece=0
                profesor_horas=profe[0]
                doc_acum=doc_acum+int(sheet_horas[doc + str(pos_num)].value)
                inv_acum=inv_acum+int(sheet_horas[inv + str(pos_num)].value)
                ext_acum=ext_acum+int(sheet_horas[ext + str(pos_num)].value)
            pos_num += 1
            pos = docum + str(pos_num)
        if(no_aparece==1):
            par_none_doce=[profe[0],None]
            docencia.append(par_none_doce)
            par_none_inv=[profe[0],None]
            investigacion.append(par_none_inv)
            par_none_ext = [profe[0], None]
            extension.append(par_none_ext)
        else:
            par1.append(profesor_horas)
            par1.append(doc_acum)
            docencia.append(par1)
            par2.append(profesor_horas)
            par2.append(inv_acum)
            investigacion.append(par2)
            par3.append(profesor_horas)
            par3.append(ext_acum)
            extension.append(par3)
        indice += 1

    #procedimiento para sumar las horas de catedra que tienen los profesores vinculados
    wb_tip_doc = load_workbook(tipo_cc)
    sheet_tipdoc = wb_tip_doc["reporte"]
    indice = 0
    par_ced_horas_doce=[]
    par_ced_horas_ext=[]
    for profe in vinculados_dan_catedra:
        pos_num = 2
        pos = 'f' + str(pos_num)
        horas_docencia=0
        horas_extension=0
        while (sheet_tipdoc[pos].value != None):
            if (str(sheet_tipdoc[pos].value) == profe[0]):
                if((str(sheet_tipdoc['c'+str(pos_num)].value) =='21820001') or(str(sheet_tipdoc['c'+str(pos_num)].value) =='21820002') or(str(sheet_tipdoc['c'+str(pos_num)].value) =='21840001')):
                    if(str(sheet_tipdoc['l'+str(pos_num)].value)=="PREGR"):
                        horas_docencia+=int(str(sheet_tipdoc['m'+str(pos_num)].value))
                elif( (str(sheet_tipdoc['c'+str(pos_num)].value) =='21860002') ):
                    if (str(sheet_tipdoc['l' + str(pos_num)].value) == "PREGR" or str(sheet_tipdoc['l' + str(pos_num)].value) == "EXTEN"):
                        horas_extension+=int(str(sheet_tipdoc['m'+str(pos_num)].value))
                else:
                    pass
            pos_num += 1
            pos = 'f' + str(pos_num)
        par4=[]
        par4.append(profe[0])
        par4.append(horas_docencia)
        par_ced_horas_doce.append(par4)

        par_exten=[]
        par_exten.append(profe[0])
        par_exten.append(horas_extension)
        par_ced_horas_ext.append(par_exten)
        indice += 1

    suma=0
    indice=0
    copia_docencia=docencia
    for i in copia_docencia:
        for j in par_ced_horas_doce:
            if i[0]==j[0]:
                if(i[1]!=None):
                    suma=j[1]+i[1]
                if(docencia[indice][1]!=None):
                    docencia[indice][1]=suma
        indice+=1

    suma = 0
    indice = 0
    copia_extension=extension
    for i in copia_extension:
        for j in par_ced_horas_ext:
            if i[0]==j[0]:
                if(i[1]!=None):
                    suma=j[1]+i[1]
                if(extension[indice][1]!=None):
                    extension[indice][1]=suma
        indice+=1

    #Procedimiento para registrar las horas de docencia que trabajan los que no aparecen
    #en el plan_20xxx_Derecho, pero si aparecen en horas_docentes_Derecho

    wb_discri = load_workbook(excel_discriminacion)
    sheet_discri = wb_discri["SQL_Results"]

    copia_docencia=docencia
    indice=0
    for ced in copia_docencia:
        if(ced[1]==None or (ced[1]==0 and extension[indice][1]==0 and investigacion[indice][1]==0)):
            sum_vincu_der_horas_cat = 0
            sum_vincu_der_horas_plan = 0
            pos_num = 2
            pos = 'A' + str(pos_num)
            while (sheet_discri[pos].value != None):
                if (str(sheet_discri[pos].value) == str(anio) + str(semest)):
                    if (ced[0] == str(sheet_discri["P" + str(pos_num)].value)):
                        if (str(sheet_discri["M" + str(pos_num)].value) == "DER" or str(
                                sheet_discri["M" + str(pos_num)].value) == "DEP" or str(
                                sheet_discri["M" + str(pos_num)].value) == "DEI"):
                            if (str(sheet_discri["S" + str(pos_num)].value) != ""):
                                sum_vincu_der_horas_cat += int(sheet_discri["S" + str(pos_num)].value)
                            if (str(sheet_discri["T" + str(pos_num)].value) != ""):
                                sum_vincu_der_horas_plan += int(sheet_discri["T" + str(pos_num)].value)
                pos_num += 1
                pos = 'A' + str(pos_num)
            #Horas de catedra de los profesores vinculados DERECHO + Horas del plan vinculados DERECHO
            docencia[indice][1]=sum_vincu_der_horas_plan+sum_vincu_der_horas_cat
        indice+=1

    # Procedimiento para discriminar las horas de aquellos profesores que tienen horas
    # en el pregrado de derecho y el de Ciencias políticas.

    docencia_vinculados = docencia
    for i in range(0, len(docencia_vinculados)):
        for j in range(0, len(vincu_der_horas_plan)):
            if (docencia_vinculados[i][0] == vincu_der_horas_plan[j][0]):
                resta = docencia[i][1] - (vincu_cpa_horas_plan[j][1] + vincu_cpa_horas_cat[j][1])
                if(vincu_der_horas_plan[j][1]==0 and vincu_der_horas_cat[j][1]==0 and resta>0):
                    docencia[i][1]=resta
                else:
                    docencia[i][1] = vincu_der_horas_plan[j][1] + vincu_der_horas_cat[j][1]
                break

    #Procedimiento para guardar en el archivo resultado la informacion de los profesores vinculados
    #nombre, apellido, tipo de documento, nivel academico, titulo obtenido, area del conocimiento e institucion donde realizo los estudios
    wb = load_workbook(archivo_resul)
    sheet = wb["profesores"]
    inicio=len(lista_pares_nom_ced)-2
    fin=len(lista_pares_nom_ced)-2+len(res)
    for i in range(inicio,fin):
        z = 'a' + str(i + 4)
        sheet[z] = i+3

        z='i'+ str(i + 4)
        sheet[z]=res[i-inicio][5]

        z = 'g' + str(i + 4)
        sheet[z] = res[i-inicio][6]

        z = 'h' + str(i + 4)
        sheet[z] = res[i-inicio][4]

        z='e'+str(i+4)
        sheet[z]=res[i-inicio][0]

        h = 'd' + str(i + 4)
        sheet[h] = res[i-inicio][2]

        x='c'+str(i+4)
        sheet[x]=apellidos[i-inicio]

        y='b'+str(i+4)
        sheet[y] = nombres[i-inicio]

        y = 'f' + str(i + 4)
        sheet[y] = res[i-inicio][3]

        y = 'j' + str(i + 4)
        e = 'k' + str(i + 4)
        if(dedicacion[i-inicio]==1):
            sheet[y]="TIEMPO COMPLETO"
            sheet[e]=900
        else:
            sheet[y] = "MEDIO TIEMPO"
            sheet[e] = 450
        if(mes_anio[i-inicio]==0):
            y='p'+str(i+4)
            sheet[y]="MESES"
        else:
            y = 'p' + str(i + 4)
            sheet[y]="AÑOS"

        y = "q" + str(i + 4)
        sheet[y] = dias_vinculacion[i-inicio]

        y = "l" + str(i + 4)
        if (docencia[i - inicio][1]==None):
            sheet[y] = " "
        else:
            sheet[y] = docencia[i - inicio][1]

        y = 'o' + str(i + 4)
        m = 'p' + str(i + 4)
        n = 'q' + str(i + 4)
        if(clase_empleado[i-inicio]=="OCASI"):
            sheet[y] = "TERMINO FIJO"
            sheet[m] = "AÑOS"
            sheet[n] = 1
        else:
            sheet[y] = "TERMINO INDEFINIDO"

        y = "m" + str(i + 4)

        if (investigacion[i - inicio][1] == None):
            sheet[y]=" "
        else:
            sheet[y] = investigacion[i - inicio][1]

        y = "n" + str(i + 4)

        if (extension[i - inicio][1] == None):
            sheet[y]=" "
        else:
            sheet[y] = extension[i - inicio][1]

    wb.save(archivo_resul)

    #Procedimiento para hallar los profesores que aparecen en la programacion pero que no
    #aparecen en los archivos de titulos.
    profes_de_derecho=lista_pares_nom_ced_profe_vincu+lista_pares_nom_ced

    profes_noregis_titu=[]
    for i in cedula_dere:
        cont_igu = 0
        for j in profes_de_derecho:
            if(i==j[0]):
                cont_igu+=1
        if(cont_igu==0):
            profes_noregis_titu.append(i)

    linea_con_profe_no_regi = []

    for j in profes_noregis_titu:
        for i in lineas_cedulas:
            if (i.find(j) != -1):
                linea_con_profe_no_regi.append(i)
                break
    nombre_noregis_titu = []
    for i in range(0, len(profes_noregis_titu)):
        inicio_nombre = (linea_con_profe_no_regi[i].find(profes_noregis_titu[i]) + len(profes_noregis_titu[i]) + 1)
        fin_nombre = len(linea_con_profe_no_regi[i]) - 1
        nombre_noregis_titu.append(linea_con_profe_no_regi[i][inicio_nombre:fin_nombre])

    wb = load_workbook(archivo_resul)
    sheet = wb["profesores"]

    apellidos = []
    nombres = []

    #Procedimiento para separar nombres de apellidos para los profesores que aparecen en
    #la programacion pero que no estan en los documentos de titulos
    for f in range(len(nombre_noregis_titu)):
        noms_apelli = nombre_noregis_titu[f].split(" ")
        space = []
        for k in noms_apelli:
            if (k != ''):
                space.append(k)
        noms_apelli = space
        if (("CASTI" in noms_apelli) and ("LONGAS" in noms_apelli) and ("ANDRES" in noms_apelli)):
            apellidos.append("DIAZ DEL CASTILLO LONGAS")
            nombres.append("ANDRES")
        elif (("D" in noms_apelli) and ("DUQUE" in noms_apelli) and ("CAMILO" in noms_apelli)):
            apellidos.append("ARANGO DUQUE")
            nombres.append("CAMILO")
        elif (len(noms_apelli) == 3):
            apellidos.append(noms_apelli[0] + " " + noms_apelli[1])
            nombres.append(noms_apelli[2])
        elif (len(noms_apelli) == 2):
            apellidos.append(noms_apelli[0])
            nombres.append(noms_apelli[1])
        elif (len(noms_apelli) == 4):
            if (noms_apelli[1] == "DE" or noms_apelli[1] == "de"):
                apellidos.append(noms_apelli[0] + " " + noms_apelli[1] + " " + noms_apelli[2])
                nombres.append(noms_apelli[3])
            else:
                apellidos.append(noms_apelli[0] + " " + noms_apelli[1])
                nombres.append(noms_apelli[2] + " " + noms_apelli[3])
        elif (len(noms_apelli) == 1):
            nombres.append(noms_apelli[0])
        elif (len(noms_apelli) >= 5):
            nombres_p = ""
            if (noms_apelli[1] == "DE" or noms_apelli[1] == "de"):
                apellidos.append(noms_apelli[0] + " " + noms_apelli[1] + " " + noms_apelli[2])
                for i in range(3, len(noms_apelli)):
                    nombres_p = nombres_p + noms_apelli[i] + " "
                nombres.append(nombres_p)
            else:
                apellidos.append(noms_apelli[0] + " " + noms_apelli[1])
                nombres_p = ""
                for i in range(2, len(noms_apelli)):
                    nombres_p = nombres_p + noms_apelli[i] + " "
                nombres.append(nombres_p)
        else:
            pass

    def buscar_docencia_horas_docentes_derecho(cedulas,anio,semest):
        wb_discri = load_workbook(excel_discriminacion)
        sheet_discri = wb_discri["SQL_Results"]
        indice = 0
        docencia=[]
        for ced in cedulas:
            sum_vincu_der_horas_cat = 0
            sum_vincu_der_horas_plan = 0
            pos_num = 2
            pos = 'A' + str(pos_num)
            while (sheet_discri[pos].value != None):
                if (str(sheet_discri[pos].value) == str(anio) + str(semest)):
                    if (ced == str(sheet_discri["P" + str(pos_num)].value)):
                        if (str(sheet_discri["M" + str(pos_num)].value) == "DER" or str(
                                sheet_discri["M" + str(pos_num)].value) == "DEP" or str(
                            sheet_discri["M" + str(pos_num)].value) == "DEI"):
                            if (str(sheet_discri["S" + str(pos_num)].value) != ""):
                                sum_vincu_der_horas_cat += int(sheet_discri["S" + str(pos_num)].value)
                            if (str(sheet_discri["T" + str(pos_num)].value) != ""):
                                sum_vincu_der_horas_plan += int(sheet_discri["T" + str(pos_num)].value)
                pos_num += 1
                pos = 'A' + str(pos_num)
            # Horas de catedra de los profesores vinculados DERECHO + Horas del plan vinculados DERECHO
            docencia.append(sum_vincu_der_horas_plan + sum_vincu_der_horas_cat)
            indice += 1
        return docencia

    horas_docencia_no_titulo=buscar_docencia_horas_docentes_derecho(profes_noregis_titu,anio,semest)

    #Procedimiento para almacenar la informacion de los profesores que aparecen en la programacion pero no en los documentos
    #de titulos: nombres, apellidos y numero de cedula

    inicio = len(lista_pares_nom_ced) + len(lista_pares_nom_ced_profe_vincu) + len(lista_pares_nom_ced_prof_institu)
    final = len(lista_pares_nom_ced) + len(lista_pares_nom_ced_profe_vincu) + len(linea_con_profe_no_regi) + len(
        lista_pares_nom_ced_prof_institu)
    for i in range(inicio, final):
        z = 'a' + str(i + 2)
        sheet[z] = i + 1

        z = 'b' + str(i + 2)
        sheet[z] = nombres[i - inicio]

        z = 'c' + str(i + 2)
        sheet[z] = apellidos[i - inicio]

        z = 'e' + str(i + 2)
        sheet[z] = profes_noregis_titu[i - inicio]

        z = 'l' + str(i + 2)
        sheet[z] = horas_docencia_no_titulo[i - inicio]

    wb.save(archivo_resul)

    print("acabe derecho")